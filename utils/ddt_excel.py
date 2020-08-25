from openpyxl import load_workbook


class ParseExcel():
    def __init__(self, excelPath, sheetName):
        self.wb = load_workbook(excelPath)
        # self.wb.active 用于载入sheet
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        dataList = []
        for line in self.sheet:
            tmpList = []
            tmpList.append(line[0].value)
            tmpList.append(line[1].value)
            dataList.append(tmpList)
        return dataList[1:]   # 从1开始，就没有读标题


if __name__ == '__main__':
    excelPath = r'E:\软件测试\问答平台\data\denglu.xlsx'
    sheetName = 'Wecenter登录平台'
    pe = ParseExcel(excelPath, sheetName)
    print(pe.getDatasFromSheet())